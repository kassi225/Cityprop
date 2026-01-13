# 1. Bibliothèques Python Standard
import datetime
from io import BytesIO
from datetime import timedelta

# 2. Bibliothèques Tierces (Data & Excel)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from num2words import num2words
from xhtml2pdf import pisa
from django.db.models import Q
from datetime import date, timedelta

# 3. Django Core (Raccourcis, Réponses & Auth)
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.template.loader import get_template
from django.urls import reverse
from django.utils import timezone
from datetime import datetime, timedelta, time

# 4. Django Database & Forms
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from django.forms import DateField, modelformset_factory
from django.shortcuts import render
from django.db.models import Q, F
from django.db.models.functions import Coalesce,Cast,TruncDate
from django.core.paginator import Paginator
from .models import Commande

# 5. Vos Modèles Locaux (Regroupés proprement)
from .models import (
    Commande, 
    CityClimaDetails, 
    TapisDetails, 
    FidelisationNote, 
    Facture, 
    FactureLigne, 
    OperationCaisse,
    TapisAlerteCommentaire  # Ajouté car présent dans votre models.py précédent
)
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from .models import Commande, CityClimaDetails, TapisDetails



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user and user.is_active:
            login(request, user)
            return redirect('dashboard')
        return render(request, 'index/login.html', {'error': 'Identifiants incorrects'})
    return render(request, 'index/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def is_admin(user):
    return user.is_superuser or user.groups.filter(name='ADMIN').exists()

@login_required
def supprimer_commande(request, id):
    if request.method == 'POST':
        commande = get_object_or_404(Commande, id=id)
        nom_client = commande.nom_client
        commande.delete()
        messages.success(request, f"La commande de {nom_client} a été supprimée avec succès.")
    return redirect('liste_fiches')


def voir_facture(request, facture_id):
    facture = get_object_or_404(Facture, id=facture_id)
    return render(request, "index/voir_facture.html", {"facture": facture})

@login_required
def supprimer_facture(request, pk):
    """ Gère la suppression d'une facture """
    facture = get_object_or_404(Facture, pk=pk)
    
    # On récupère l'ID de la commande AVANT de supprimer la facture
    # car votre modèle utilise l'attribut 'commande'
    fiche_id = facture.commande.id 

    if request.method == "POST":
        facture.delete()
        messages.success(request, "Le document a été supprimé avec succès.")
        
        # Redirection vers la fiche avec le bon nom de paramètre 'fiche_id'
        return redirect('detail_fiche', fiche_id=fiche_id)

    # Si on arrive ici sans POST (sécurité), on redirige simplement
    return redirect('detail_fiche', fiche_id=fiche_id)
@login_required
@transaction.atomic
def modifier_facture(request, pk):
    """ Gère la modification d'une facture existante """
    facture = get_object_or_404(Facture, pk=pk)
    commande = facture.commande 

    if request.method == "POST":
        try:
            
            # 1. Mise à jour des infos de base
            facture.type_document = request.POST.get('type_document')
            facture.objet = request.POST.get('objet')
            facture.date_emission = request.POST.get('date_emission')
            
            # Conversion sécurisée du taux
            taux_post = request.POST.get('taux_reduction') or 0
            facture.taux_reduction_pourcentage = float(str(taux_post).replace(',', '.'))
            facture.save()

            # 2. Mise à jour des lignes
            facture.lignes.all().delete() 
            
            designations = request.POST.getlist('designation[]')
            quantites = request.POST.getlist('quantite[]')
            prix_unitaires = request.POST.getlist('prix_unitaire[]')
            notes = request.POST.getlist('note_prix[]')

            from .models import FactureLigne
            for i in range(len(designations)):
                if designations[i].strip():
                    qty = int(float(quantites[i] or 0))
                    pu = int(float(prix_unitaires[i] or 0))
                    
                    FactureLigne.objects.create(
                        facture=facture,
                        designation=designations[i],
                        quantite=qty,
                        prix_unitaire=pu,
                        note_prix_unitaire=notes[i]
                    )
            
            # 3. Recalcul final
            facture.update_final_amount()
            facture.save()

            messages.success(request, "Le document a été modifié avec succès.")
            return redirect('detail_fiche', fiche_id=commande.id)

        except Exception as e:
            # C'est cette partie qui manquait ou était mal placée
            messages.error(request, f"Erreur lors de la modification : {e}")
            return redirect('modifier_facture', pk=pk)

    return render(request, 'index/modifier_facture.html', {
        'facture': facture,
        'commande': commande
    })

# Cette fonction est essentielle pour la conversion PDF

def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html  = template.render(context_dict)
    
    # Créer un tampon binaire en mémoire (BytesIO)
    result = BytesIO()
    
    # Générer le PDF (en utilisant le template facture_detail.html)
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    # Pour le débogage, vous pouvez retourner le HTML si erreur
    return HttpResponse('Nous avons rencontré des erreurs lors de la génération du PDF: %s' % html)


def telecharger_devis_pdf(request, facture_id):
    """
    Génère et télécharge le PDF d'une facture/devis spécifique.
    """
    facture = get_object_or_404(Facture, id=facture_id) 
    
    # 🔑 CORRECTION CRITIQUE : Ajout de l'objet 'request' au contexte
    # Ceci permet au template de construire l'URL absolue du logo.
    context = {
        'facture': facture,
        'request': request,  
    }

    # Nom du fichier pour le téléchargement
    filename = f'{facture.type_document}_CityProp_{facture.numero_document}.pdf'
    
    # Le template pour le PDF est 'facture_telechargement.html'
    # Assurez-vous que render_to_pdf utilise le moteur de template Django correctement
    pdf_response = render_to_pdf('index/facture_telechargement.html', context)
    
    if pdf_response:
        # Indique au navigateur de télécharger le fichier (attachment)
        pdf_response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return pdf_response
    
    return HttpResponse("Impossible de générer le PDF.")

# --- Vues de Facturation (corrigée avec la logique robuste) ---

@login_required
def dashboard(request):
    # Temps présent
    now = timezone.now()
    today = now.date()
    
    # 1. RÉCUPÉRATION DES FILTRES (Dates exactes)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    commande_filter = Q()
    details_filter = Q()

    if start_date_str and end_date_str:
        try:
            # Conversion des strings en objets date
            d_start = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            d_end = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            
            # Création de datetimes aware pour couvrir toute la journée (00:00:00 à 23:59:59)
            dt_start = timezone.make_aware(datetime.combine(d_start, time.min))
            dt_end = timezone.make_aware(datetime.combine(d_end, time.max))
            
            commande_filter &= Q(date_creation__range=(dt_start, dt_end))
            details_filter &= Q(commande__date_creation__range=(dt_start, dt_end))
        except ValueError:
            pass

    # 2. COMPTAGES (Basés sur les filtres)
    total_commandes = Commande.objects.filter(commande_filter).count()
    total_cityprop = Commande.objects.filter(commande_filter, type_commande="CITYPROP").count()
    total_clim = Commande.objects.filter(commande_filter, type_commande="CLIMATISEUR").count()
    total_tapis = Commande.objects.filter(commande_filter, type_commande="TAPISPROP").count()

    # 3. FIDÉLISATION
    fidelises = CityClimaDetails.objects.filter(details_filter, fidelise=True).count() + \
                TapisDetails.objects.filter(details_filter, fidelise=True).count()
    non_fidelises = total_commandes - fidelises
    
    # 4. ALERTES (Filtrées par période d'entrée)
    alertes_tapis_7j = TapisDetails.objects.filter(
        details_filter,
        date_ramassage__lte=today - timedelta(days=11),
        statut__in=["NON_RESPECTE","PRET","CLIENT_INDISPO"]
    ).count()

    alertes_city = CityClimaDetails.objects.filter(details_filter, fidelise=False, commande__type_commande='CITYPROP', date_intervention__lte=today - timedelta(days=180)).count()
    alertes_clima = CityClimaDetails.objects.filter(details_filter, fidelise=False, commande__type_commande='CLIMATISEUR', date_intervention__lte=today - timedelta(days=90)).count()
    alertes_tapis_fidelisation = TapisDetails.objects.filter(details_filter, fidelise=False, date_livraison__lte=today - timedelta(days=180)).count()

    # 5. LISTES ET TOP CLIENTS
    commandes_recents = Commande.objects.filter(commande_filter).order_by('-date_creation')[:5]
    top_clients = Commande.objects.filter(commande_filter).values("nom_client").annotate(total=Count("id")).order_by("-total")[:5]

    context = {
        "total_commandes": total_commandes,
        "total_cityprop": total_cityprop,
        "total_clim": total_clim,
        "total_tapis": total_tapis,
        "fidelises": fidelises,
        "non_fidelises": non_fidelises,
        "alertes_city": alertes_city,
        "alertes_clima": alertes_clima,
        "alertes_tapis_fidelisation": alertes_tapis_fidelisation,
        "alertes_tapis_7j": alertes_tapis_7j,
        "commandes_recents": commandes_recents,
        "top_clients": top_clients,
        "start_date": start_date_str,
        "end_date": end_date_str,
    }

    return render(request, "index/dashboard.html", context)


@login_required
def nouvelle_commande(request):
    if request.method == 'POST':
        nom = request.POST.get('nom_client')
        numero = request.POST.get('numero_client')
        loc = request.POST.get('localisation_client')
        type_cmd = request.POST.get('type_commande')

        if not nom or not numero or not loc or not type_cmd:
            messages.error(request, "Veuillez remplir tous les champs obligatoires.")
            return render(request, 'index/nouvelle_commande.html')

        # 1. Création de la commande de base
        commande = Commande.objects.create(
            nom_client=nom,
            numero_client=numero,
            localisation_client=loc,
            type_commande=type_cmd
        )

        # 2. Section CITYCLIMA (Uniquement si le type correspond)
        if type_cmd in ['CITYPROP', 'CLIMATISEUR']:
            date_inter = request.POST.get('date_intervention')
            satisfaction = request.POST.get('satisfaction')
            designation = request.POST.get('designation')
            cout_clim = request.POST.get('cout_clim')

            if any([date_inter, satisfaction, designation, cout_clim]):
                CityClimaDetails.objects.create(
                    commande=commande,
                    date_intervention=date_inter or None,
                    satisfaction=satisfaction or None,
                    designation=designation or "",
                    cout=int(cout_clim) if cout_clim else 0  # Conversion en entier
                )

        # 3. Section TAPIS (Uniquement si le type correspond)
        elif type_cmd == 'TAPISPROP':
            date_ramassage = request.POST.get('date_ramassage')
            nombre = request.POST.get('nombre_tapis')
            cout_tapis = request.POST.get('cout')
            traitement = request.POST.get('date_traitement')
            prevue = request.POST.get('date_prevue_livraison')
            livraison = request.POST.get('date_livraison')
            commentaire = request.POST.get('commentaire')
            statut = request.POST.get('statut')

            if any([date_ramassage, nombre, cout_tapis, traitement, prevue, livraison, commentaire]):
                TapisDetails.objects.create(
                    commande=commande,
                    date_ramassage=date_ramassage or None,
                    nombre_tapis=int(nombre) if nombre else 0,
                    cout=int(cout_tapis) if cout_tapis else 0,
                    date_traitement=traitement or None,
                    date_prevue_livraison=prevue or None,
                    date_livraison=livraison or None,
                    commentaire=commentaire or "",
                    statut=statut or "NON_RESPECTE"
                )

        messages.success(request, "La commande a été créée avec succès !")
        return redirect('liste_fiches')

    return render(request, 'index/nouvelle_commande.html')
@login_required
def liste_fiches(request):
    # 1. Optimisation SQL initiale
    commandes_qs = Commande.objects.select_related('cityclimadetails', 'tapisdetails').all()

    # 2. Récupération des filtres
    search_query = request.GET.get('q', '')
    type_filter = request.GET.get('type_commande', '')
    statut_filter = request.GET.get('statut', '')
    nom_filter = request.GET.get('nom_client', '')
    numero_filter = request.GET.get('numero_client', '')
    date_crea = request.GET.get('date_crea', '') 
    date_debut = request.GET.get('date_debut', '') 
    date_fin = request.GET.get('date_fin', '')     
    fidelise_filter = request.GET.get('fidelise', '')

    # 3. Application des filtres SQL
    if search_query:
        commandes_qs = commandes_qs.filter(
            Q(nom_client__icontains=search_query) |
            Q(numero_client__icontains=search_query) |
            Q(localisation_client__icontains=search_query)
        )
    if type_filter:
        commandes_qs = commandes_qs.filter(type_commande=type_filter)
    
    if nom_filter:
        commandes_qs = commandes_qs.filter(nom_client__icontains=nom_filter)
        
    if numero_filter:
        commandes_qs = commandes_qs.filter(numero_client__icontains=numero_filter)

    if date_crea:
        commandes_qs = commandes_qs.filter(date_creation__date=date_crea)

    if statut_filter:
        commandes_qs = commandes_qs.filter(
            Q(tapisdetails__statut=statut_filter) | 
            Q(cityclimadetails__satisfaction=statut_filter)
        )

    if fidelise_filter == "oui":
        commandes_qs = commandes_qs.filter(Q(cityclimadetails__fidelise=True) | Q(tapisdetails__fidelise=True))
    elif fidelise_filter == "non":
        commandes_qs = commandes_qs.filter(Q(cityclimadetails__fidelise=False) | Q(tapisdetails__fidelise=False))

    # 4. Conversion en liste pour le TRI MÉTIER (Dates hybrides)
    commandes_list = list(commandes_qs.distinct())

    def obtenir_date_tri(cmd):
        """Récupère la date d'opération selon le type"""
        try:
            if cmd.type_commande == 'TAPISPROP' and hasattr(cmd, 'tapisdetails') and cmd.tapisdetails:
                return cmd.tapisdetails.date_ramassage or cmd.date_creation.date()
            if hasattr(cmd, 'cityclimadetails') and cmd.cityclimadetails:
                return cmd.cityclimadetails.date_intervention or cmd.date_creation.date()
        except:
            pass
        return cmd.date_creation.date()

    # Tri : Plus récent au plus ancien
    commandes_list.sort(key=obtenir_date_tri, reverse=True)

    # 5. Filtrage par PLAGE (Correction de l'AttributeError ici)
    if date_debut or date_fin:
        try:
            # On utilise datetime.strptime car "from datetime import datetime" est utilisé
            d_deb = datetime.strptime(date_debut, '%Y-%m-%d').date() if date_debut else None
            d_fin = datetime.strptime(date_fin, '%Y-%m-%d').date() if date_fin else None
            
            commandes_list = [
                c for c in commandes_list 
                if (not d_deb or obtenir_date_tri(c) >= d_deb) and (not d_fin or obtenir_date_tri(c) <= d_fin)
            ]
        except (ValueError, TypeError):
            pass

    # 6. Pagination
    paginator = Paginator(commandes_list, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 7. Contexte
    context = {
        'commandes': page_obj,
        'today': timezone.now().date(),
        'search_query': search_query,
        'type_filter': type_filter,
        'statut_filter': statut_filter,
        'nom_filter': nom_filter,
        'numero_filter': numero_filter,
        'date_crea': date_crea,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'fidelise_filter': fidelise_filter,
    }
    
    return render(request, 'index/liste_fiches.html', context)



@login_required
def detail_fiche(request, fiche_id):
    commande = get_object_or_404(Commande, id=fiche_id)
    
    # Récupération des détails existants via les related_names
    cityclima = getattr(commande, 'cityclimadetails', None)
    tapis = getattr(commande, 'tapisdetails', None)

    if request.method == 'POST':
        # --- 1. Mise à jour champs généraux ---
        commande.nom_client = request.POST.get('nom_client', commande.nom_client)
        commande.numero_client = request.POST.get('numero_client', commande.numero_client)
        commande.localisation_client = request.POST.get('localisation_client', commande.localisation_client)
        commande.type_commande = request.POST.get('type_commande', commande.type_commande)
        commande.save()

        # --- 2. Mise à jour détails CityClima (Désignation + Coût inclus) ---
        if commande.type_commande in ['CITYPROP', 'CLIMATISEUR']:
            date_int = request.POST.get('date_intervention') or None
            satisfact = request.POST.get('satisfaction') or None
            desig = request.POST.get('designation', '')
            c_clim = request.POST.get('cout_clim') or 0
            
            if cityclima:
                cityclima.date_intervention = date_int
                cityclima.satisfaction = satisfact
                cityclima.designation = desig
                cityclima.cout = int(c_clim)
                cityclima.save()
            elif any([date_int, satisfact, desig, c_clim]):
                CityClimaDetails.objects.create(
                    commande=commande,
                    date_intervention=date_int,
                    satisfaction=satisfact,
                    designation=desig,
                    cout=int(c_clim)
                )

        # --- 3. Mise à jour détails Tapis (Date prévue incluse) ---
        elif commande.type_commande == 'TAPISPROP':
            d_ramassage = request.POST.get('date_ramassage') or None
            d_traitement = request.POST.get('date_traitement') or None
            d_prevue = request.POST.get('date_prevue_livraison') or None # Nouveau
            d_livraison = request.POST.get('date_livraison') or None
            nb_tapis = request.POST.get('nombre_tapis') or 0
            prix_cout = request.POST.get('cout') or 0
            stat = request.POST.get('statut') or 'NON_RESPECTE'
            comm = request.POST.get('commentaire', '')

            if tapis:
                tapis.date_ramassage = d_ramassage
                tapis.date_traitement = d_traitement
                tapis.date_prevue_livraison = d_prevue # Mise à jour
                tapis.date_livraison = d_livraison
                tapis.nombre_tapis = int(nb_tapis)
                tapis.cout = int(prix_cout)
                tapis.statut = stat
                tapis.commentaire = comm
                tapis.save()
            else:
                if any([d_ramassage, d_traitement, d_prevue, d_livraison, nb_tapis, prix_cout]):
                    TapisDetails.objects.create(
                        commande=commande,
                        date_ramassage=d_ramassage,
                        date_traitement=d_traitement,
                        date_prevue_livraison=d_prevue, # Création
                        date_livraison=d_livraison,
                        nombre_tapis=int(nb_tapis),
                        cout=int(prix_cout),
                        statut=stat,
                        commentaire=comm
                    )

        messages.success(request, "La fiche a été mise à jour avec succès !")
        return redirect('detail_fiche', fiche_id=fiche_id)

    # Contexte pour le rendu
    context = {
        'commande': commande,
        'cityclima': cityclima,
        'tapis': tapis,
        'factures': Facture.objects.filter(commande=commande).order_by('-date_emission'),
    }
    return render(request, 'index/detail_fiche.html', context)
@login_required
def creer_facture(request, fiche_id):
    """
    Crée une nouvelle Facture et ses lignes associées à partir du formulaire.
    """
    if request.method != 'POST':
        return redirect('detail_fiche', fiche_id=fiche_id)

    commande = get_object_or_404(Commande, id=fiche_id)
    
    try:
        with transaction.atomic():
            
            # 1. Création de la Facture principale
            facture = Facture.objects.create(
                commande=commande,
                type_document=request.POST.get('type_document', 'DEVIS'),
                objet=request.POST.get('objet', ''),
                lieu_emission=request.POST.get('lieu_emission', 'Abidjan'),
                signature=request.POST.get('signature', ''),
                # Le numero_document est généré dans Facture.save()
            )

            # 2. Récupération et préparation des lignes
            designations = request.POST.getlist('designation[]')
            quantites = request.POST.getlist('quantite[]')
            prix_unitaires = request.POST.getlist('prix_unitaire[]')
            notes_prix = request.POST.getlist('note_prix[]')
            
            lignes_a_creer = []
            lignes_valides_trouvees = False

            # Parcourir et valider les lignes
            for i in range(len(designations)):
                designation = designations[i].strip()
                
                if not designation:
                    continue

                try:
                    # Conversion des valeurs numériques
                    quantite = int(quantites[i]) if quantites[i] else 0
                    prix_unitaire = int(prix_unitaires[i]) if prix_unitaires[i] else 0
                    note_prix = notes_prix[i].strip() if i < len(notes_prix) else ''
                    
                    if quantite > 0 and prix_unitaire > 0:
                        lignes_valides_trouvees = True
                        lignes_a_creer.append(FactureLigne(
                            facture=facture,
                            designation=designation,
                            quantite=quantite,
                            prix_unitaire=prix_unitaire,
                            note_prix_unitaire=note_prix
                        ))
                    
                except (ValueError, IndexError):
                    # Lever une erreur qui sera capturée par le bloc try/except externe
                    raise ValueError(f"Les valeurs de quantité ou de prix unitaire de la ligne '{designation}' sont incorrectes.")

            # 3. Création en masse des lignes
            if lignes_a_creer:
                FactureLigne.objects.bulk_create(lignes_a_creer)
            else:
                messages.warning(request, "La facture a été créée sans ligne d'article (au moins une désignation, quantité et prix unitaire non nuls sont requis).")

            messages.success(request, f"{facture.type_document} N°{facture.numero_document} créé avec succès!")

            # Rediriger vers la page d'affichage HTML de la facture
            return redirect('index/voir_facture', facture_id=facture.id) 
            
    except ValueError as e:
        messages.error(request, f"Erreur lors de la création du document: {e}")
        return redirect('detail_fiche', fiche_id=fiche_id)
    except Exception as e:
        messages.error(request, f"Une erreur inattendue est survenue: {e}")
        return redirect('detail_fiche', fiche_id=fiche_id)

def detail_facture(request, facture_id):
    facture = get_object_or_404(Facture, id=facture_id)
    return render(request, 'index/facture_telecharger.html', {
        'facture': facture
    })

def voir_facture(request, facture_id):
    """Affiche la facture dans une page Web (HTML)."""
    facture = get_object_or_404(Facture, id=facture_id)
    return render(request, "index/voir_facture.html", {"facture": facture})


@login_required
def alertes_fidelisation(request):
    now = timezone.now().date()
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Construction du filtre de recherche
    search_filter = Q()
    if search_query:
        search_filter = (Q(commande__nom_client__icontains=search_query) | 
                         Q(commande__numero_client__icontains=search_query))
                        
    # 1. Récupération CITYPROP (> 180 jours)
    city_qs = CityClimaDetails.objects.filter(
        search_filter, 
        fidelise=False,
        commande__type_commande='CITYPROP',
        date_intervention__lte=now - timedelta(days=180)
    ).select_related('commande')

    # 2. Récupération CLIMATISEUR (> 90 jours)
    clima_qs = CityClimaDetails.objects.filter(
        search_filter, 
        fidelise=False,
        commande__type_commande='CLIMATISEUR',
        date_intervention__lte=now - timedelta(days=90)
    ).select_related('commande')

    # 3. Récupération TAPIS (> 180 jours après LIVRAISON)
    tapis_qs = TapisDetails.objects.filter(
        search_filter, 
        fidelise=False,
        statut__in=['LIVRE_SATISFAIT', 'LIVRE_INSATISFAIT'],
        date_livraison__lte=now - timedelta(days=180)
    ).select_related('commande')

    alertes_list = []
    
    # Normalisation
    for obj in city_qs:
        alertes_list.append({
            'id': obj.id,
            'type': "CITYPROP",
            'nom_client': obj.commande.nom_client,
            'numero_client': obj.commande.numero_client,
            'date_cle': obj.date_intervention,
            'url_commande': obj.commande.id,
        })

    for obj in clima_qs:
        alertes_list.append({
            'id': obj.id,
            'type': "CLIMATISEUR",
            'nom_client': obj.commande.nom_client,
            'numero_client': obj.commande.numero_client,
            'date_cle': obj.date_intervention,
            'url_commande': obj.commande.id,
        })
        
    for obj in tapis_qs:
        alertes_list.append({
            'id': obj.id,
            'type': "TAPISPROP",
            'nom_client': obj.commande.nom_client,
            'numero_client': obj.commande.numero_client,
            'date_cle': obj.date_livraison, 
            'url_commande': obj.commande.id,
        })
        
    # Tri par date (la plus ancienne en haut)
    alertes_list.sort(key=lambda x: x['date_cle'] if x['date_cle'] else now)
    
    # PAGINATION MISE À JOUR : 7 éléments par page
    paginator = Paginator(alertes_list, 7) 
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "total_alertes": len(alertes_list),
        "search_query": search_query,
    }
    return render(request, "index/fidelisation.html", context)


@login_required
def marquer_fidelise(request, id):
    """
    Marque une commande comme fidélisée et enregistre un commentaire facultatif.
    """
    
    if request.method != 'POST':
        return redirect('alertes_fidelisation')

    commentaire = request.POST.get("commentaire", "").strip()
    # fidelise_check est 'True' uniquement si la case est cochée (pour les cas non encore fidélisés)
    fidelise_action = request.POST.get("fidelise_check") == 'True'
    
    try:
        # 1. Identifier le détail
        detail = None
        if CityClimaDetails.objects.filter(id=id).exists():
            detail = CityClimaDetails.objects.get(id=id)
            type_detail = "CITYCLIMA"
        elif TapisDetails.objects.filter(id=id).exists():
            detail = TapisDetails.objects.get(id=id)
            type_detail = "TAPIS"
        else:
            messages.error(request, "Détail de commande introuvable.")
            return redirect('alertes_fidelisation')

        # 2. Logique de Marquage (seulement si non encore fidélisé ET case cochée)
        marquage_effectue = False
        if fidelise_action and not detail.fidelise:
            detail.fidelise = True
            detail.save()
            marquage_effectue = True
            # Message de succès pour le marquage de fidélisation
            messages.success(request, f"Le client {detail.commande.nom_client} est désormais FIDÉLISÉ. Il ne sera plus en alerte.")
        
        # 3. Enregistrement du commentaire (si présent)
        if commentaire:
            
            # Si le marquage a eu lieu lors de cette soumission, on le note dans la FidelisationNote
            is_fidelise_note = marquage_effectue

            # Création de la FidelisationNote
            if type_detail == "CITYCLIMA":
                FidelisationNote.objects.create(
                    detail_city_clima=detail,
                    commentaire=commentaire,
                    fidelise_marquee=is_fidelise_note
                )
            else:
                FidelisationNote.objects.create(
                    detail_tapis=detail,
                    commentaire=commentaire,
                    fidelise_marquee=is_fidelise_note
                )
            
            # Message de succès pour l'ajout de commentaire seul (si pas de marquage)
            if not marquage_effectue:
                 messages.success(request, "Observation ajoutée avec succès.")


    except Exception as e:
        print("Erreur fidélisation:", e)
        messages.error(request, "Impossible de traiter la demande de fidélisation.")

    # Rediriger vers la page détaillée pour voir le nouvel état/commentaire
    return redirect('detail_fidelisation', id=id)

@login_required
def detail_fidelisation(request, id):
    """
    Page pour gérer la fidélisation d'une commande.
    Affiche les infos client et permet d'ajouter un commentaire de négociation.
    """
    try:
        if CityClimaDetails.objects.filter(id=id).exists():
            detail = CityClimaDetails.objects.get(id=id)
            type_detail = "CITYCLIMA"
        else:
            detail = TapisDetails.objects.get(id=id)
            type_detail = "TAPIS"
    except:
        messages.error(request, "Détail introuvable.")
        return redirect('liste_fiches')

    if request.method == "POST":
        commentaire = request.POST.get("commentaire", "").strip()
        fidelise = request.POST.get("fidelise") == "on"

        # Sauvegarder la note si commentaire
        if commentaire:
            if type_detail == "CITYCLIMA":
                FidelisationNote.objects.create(
                    detail_city_clima=detail,
                    commentaire=commentaire
                )
            else:
                FidelisationNote.objects.create(
                    detail_tapis=detail,
                    commentaire=commentaire
                )

        # Marquer fidélisé si coché
        if fidelise:
            detail.fidelise = True
            detail.save()
            messages.success(request, "Client fidélisé avec succès !")
            return redirect('liste_fiches')

        messages.success(request, "Commentaire enregistré.")
        return redirect('detail_fidelisation', id=id)

    context = {
        "detail": detail,
        "type_detail": type_detail
    }
    return render(request, "index/detail_fidelisation.html", context)


@login_required
def alertes_tapis_retard(request):
    """
    Liste des alertes tapis dont le ramassage est dépassé de 7 jours,
    avec filtres et pagination.
    """

    today = timezone.now().date()

    # Statuts qui stoppent l'alerte
    STATUTS_FINAUX = [
       'LIVRE_SATISFAIT',
        'LIVRE_INSATISFAIT',
        'ABANDON'
    ]

    # =========================
    # FILTRES (GET)
    # =========================
    search = request.GET.get("search", "")
    date_ramassage = request.GET.get("date_ramassage", "")
    nb_tapis = request.GET.get("nb_tapis", "")

    # =========================
    # QUERYSET DE BASE
    # =========================
    alertes = TapisDetails.objects.filter(
        date_ramassage__isnull=False,
        date_ramassage__lte=today - timedelta(days=11)
    ).exclude(
        statut__in=STATUTS_FINAUX
    )

    # =========================
    # RECHERCHE GLOBALE
    # =========================
    if search:
        alertes = alertes.filter(
            Q(commande__nom_client__icontains=search) |
            Q(commande__numero_client__icontains=search) |
            Q(commande__localisation_client__icontains=search)
        )

    # =========================
    # FILTRES SPÉCIFIQUES
    # =========================
    if date_ramassage:
        alertes = alertes.filter(date_ramassage=date_ramassage)

    if nb_tapis:
        alertes = alertes.filter(nombre_tapis=nb_tapis)

    alertes = alertes.order_by("-date_ramassage")

    # =========================
    # PAGINATION
    # =========================
    paginator = Paginator(alertes, 10)  # 10 lignes / page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "alertes": page_obj,
        "page_obj": page_obj,
        "total_alertes": alertes.count(),

        # valeurs filtres
        "search": search,
        "date_ramassage": date_ramassage,
        "nb_tapis": nb_tapis,
    }

    return render(request, "index/alerte_tapis.html", context)



@login_required
def detail_alerte_tapis(request, id):
    tapis = get_object_or_404(TapisDetails, id=id)
    
    # 1. Détermination du contexte de l'appel
    url_name = request.resolver_match.url_name
    
    STATUTS_LIVRES = ["LIVRE-satisfait", "LIVRE-insatisfait"]
    STATUT_ABANDON = "ABANDON"
    STATUTS_FINAUX_ALERTE = STATUTS_LIVRES + [STATUT_ABANDON] # Statuts qui suppriment l'alerte de la liste initiale

    # =======================================================
    # 2. GESTION DES GARDES-FOUS (Accès en GET)
    # =======================================================

    if url_name == "detail_alerte_tapis_abandon":
        # Si nous sommes sur l'URL d'abandon :
        # On bloque uniquement si c'est LIVRÉ, car ABANDON doit rester accessible ici
        if tapis.statut in STATUTS_LIVRES:
            messages.info(request, "Cette commande a été livrée. Elle n'est plus en gestion d'abandon.")
            return redirect("alertes_tapis_retard")
        
        # Et si le statut n'est pas ABANDON, on redirige aussi pour rester propre
        elif tapis.statut != STATUT_ABANDON:
             messages.info(request, "Cette commande n'est pas abandonnée. Affichage classique des alertes.")
             return redirect("detail_alerte_tapis", id=id)

    else: # Contexte par défaut (detail_alerte_tapis)
        # Si nous sommes sur l'URL d'alerte normale :
        # On bloque si le statut est LIVRÉ OU ABANDONNÉ
        if tapis.statut in STATUTS_FINAUX_ALERTE:
            messages.info(request, "Cette commande n'est plus en alerte (livrée ou abandonnée).")
            return redirect("alertes_tapis_retard")


    # =======================================================
    # 3. GESTION DE LA SOUMISSION (POST)
    # =======================================================

    if request.method == "POST":
        # ... (Nouveau commentaire - code inchangé) ...
        texte = request.POST.get("commentaire", "").strip()
        if texte:
             from .models import TapisAlerteCommentaire
             TapisAlerteCommentaire.objects.create(
                 tapis=tapis,
                 texte=texte
             )
             messages.success(request, "Commentaire ajouté avec succès.")

        # Mise à jour du statut
        new_statut = request.POST.get("statut")
        if new_statut:
            tapis.statut = new_statut
            tapis.save()
            messages.success(request, "Statut mis à jour.")

            # Gestion de la redirection après résolution de l'alerte
            if new_statut in STATUTS_LIVRES:
                # Si LIVRÉ, l'alerte est levée et on redirige vers la liste des alertes (d'où elle disparaîtra)
                return redirect("alertes_tapis_retard")
            
            elif new_statut == STATUT_ABANDON:
                # Si ABANDONNÉ, l'alerte est levée, on redirige vers l'URL d'abandon
                # (Assurez-vous que cette URL pointe vers une VUE qui liste les commandes ABANDONNÉES)
                # Si vous n'avez pas de vue de LISTE d'abandon, vous pouvez rediriger vers la page de détail d'abandon
                return redirect("detail_alerte_tapis_abandon", id=id)

        # Redirection par défaut (si seulement un commentaire a été ajouté ou si le statut n'est ni LIVRÉ ni ABANDON)
        # On redirige vers l'URL par laquelle l'utilisateur est entré pour conserver le contexte.
        return redirect(url_name, id=id) 

    # =======================================================
    # 4. Rendu de la page (GET)
    # =======================================================
    context = {
        "tapis": tapis,
        "commentaires": tapis.commentaires_alerte.order_by("-date"),
        # Optionnel: pour adapter l'affichage du template HTML
        "is_abandon_context": url_name == "detail_alerte_tapis_abandon" 
    }

    return render(request, "index/detail_alerte_tapis.html", context)

@login_required
@login_required
def export_commandes_excel(request):
    commandes = Commande.objects.all().order_by('-date_creation')

    # ... (Gardez vos filtres existants ici sans changement) ...
    # [Filtres type_filter, nom_filter, etc.]

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes Détaillées"

    # 🎨 Styles
    header_fill = PatternFill(start_color="157347", end_color="157347", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # 📌 Nouveaux En-têtes (Tous les champs)
    headers = [
        "Client", "Numéro", "Localisation", "Type", "Date Création",
        "Désignation / Détails", "Coût (FCFA)", "Date Intervention / Ramassage",
        "Fin de Traitement", "Prévu Livraison", "Livraison Réelle",
        "Statut / Satisfaction", "Fidélisé", "Commentaire"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 📊 Données
    for c in commandes:
        desig_details = "-"
        cout = 0
        date_1 = "-" # Intervention ou Ramassage
        date_2 = "-" # Fin de Traitement (Tapis uniquement)
        date_3 = "-" # Prévue (Tapis uniquement)
        date_4 = "-" # Réelle (Tapis uniquement)
        statut_text = "-"
        fidelise = "Non"
        comm = "-"

        if c.type_commande in ["CITYPROP", "CLIMATISEUR"] and hasattr(c, "cityclimadetails"):
            d = c.cityclimadetails
            desig_details = d.designation or "Sans désignation"
            cout = d.cout or 0
            date_1 = d.date_intervention.strftime("%d/%m/%Y") if d.date_intervention else "-"
            statut_text = d.get_satisfaction_display() if d.satisfaction else "-"
            fidelise = "Oui" if d.fidelise else "Non"

        elif c.type_commande == "TAPISPROP" and hasattr(c, "tapisdetails"):
            t = c.tapisdetails
            desig_details = f"{t.nombre_tapis} tapis"
            cout = t.cout or 0
            date_1 = t.date_ramassage.strftime("%d/%m/%Y") if t.date_ramassage else "-"
            date_2 = t.date_traitement.strftime("%d/%m/%Y") if t.date_traitement else "-"
            date_3 = t.date_prevue_livraison.strftime("%d/%m/%Y") if t.date_prevue_livraison else "-"
            date_4 = t.date_livraison.strftime("%d/%m/%Y") if t.date_livraison else "-"
            statut_text = t.get_statut_display()
            fidelise = "Oui" if t.fidelise else "Non"
            comm = t.commentaire or "-"

        ws.append([
            c.nom_client,
            c.numero_client,
            c.localisation_client,
            c.type_commande,
            c.date_creation.strftime("%d/%m/%Y"),
            desig_details,
            cout,
            date_1,
            date_2,
            date_3,
            date_4,
            statut_text,
            fidelise,
            comm
        ])

    # 📐 Ajustement automatique des colonnes (Largeurs adaptées)
    widths = [20, 15, 25, 12, 12, 35, 12, 15, 15, 15, 15, 20, 10, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="export_complet_cityprop.xlsx"'
    wb.save(response)
    return response

@login_required
def alertes_counts(request):
    today = timezone.now().date()
    
    # 1. Alertes Fidélisation (Global : City, Clima et Tapis livrés depuis longtemps)
    count_city = CityClimaDetails.objects.filter(
        fidelise=False, commande__type_commande='CITYPROP', 
        date_intervention__lte=today - timedelta(days=180)).count()
    
    count_clima = CityClimaDetails.objects.filter(
        fidelise=False, commande__type_commande='CLIMATISEUR', 
        date_intervention__lte=today - timedelta(days=90)).count()
        
    count_tapis_fid = TapisDetails.objects.filter(
        fidelise=False, statut__in=['LIVRE_SATISFAIT', 'LIVRE_INSATISFAIT'],
        date_livraison__lte=today - timedelta(days=180)).count()

    # 2. Alertes Retard Tapis (Tapis ramassés mais non livrés après 7 jours)
    alertes_tapis_retard = TapisDetails.objects.filter(
        date_ramassage__isnull=False,
        date_ramassage__lte=today - timedelta(days=7)
    ).exclude(statut__in=['LIVRE_SATISFAIT', 'LIVRE_INSATISFAIT', 'ABANDON']).count()

    return {
        'total_fidelisation': count_city + count_clima + count_tapis_fid,
        'alertes_tapis_retard': alertes_tapis_retard,
    }

@login_required
def creer_facture(request, fiche_id):
    commande = get_object_or_404(Commande, id=fiche_id)
    
    if request.method == 'POST':
        # 1. Récupération sécurisée du taux de réduction
        taux_reduction_str = request.POST.get('taux_reduction', '0.00') # Vérifiez le 'name' dans votre HTML
        try:
            taux_reduction = float(taux_reduction_str.replace(',', '.'))
        except ValueError:
            taux_reduction = 0.00
            
        # 2. Création de l'objet Facture
        facture = Facture.objects.create(
            commande=commande,
            type_document=request.POST.get('type_document'),
            objet=request.POST.get('objet'),
            lieu_emission=request.POST.get('lieu_emission', 'Abidjan'),
            signature=request.POST.get('signature', 'LA COMPTABILITÉ'),
            taux_reduction_pourcentage=taux_reduction,
        )
        
        # 3. Traitement des lignes de prestation (Inclusion de la NOTE)
        designations = request.POST.getlist('designation[]')
        quantites = request.POST.getlist('quantite[]')
        prix_unitaires = request.POST.getlist('prix_unitaire[]')
        notes = request.POST.getlist('note_prix[]') # <--- RÉCUPÉRATION DES NOTES

        # On utilise zip pour parcourir toutes les listes en même temps
        for desig, qty_str, pu_str, note in zip(designations, quantites, prix_unitaires, notes):
            if desig.strip(): # On n'enregistre que si la désignation n'est pas vide
                try:
                    qty = float(qty_str.replace(',', '.'))
                    pu = float(pu_str.replace(',', '.'))
                    
                    FactureLigne.objects.create(
                        facture=facture,
                        designation=desig,
                        quantite=qty,
                        prix_unitaire=pu,
                        note_prix_unitaire=note # <--- ENREGISTREMENT DE LA NOTE
                    )
                except ValueError:
                    continue 

        # 4. Calcul final et sauvegarde
        facture.update_final_amount() 
        facture.save()

        return redirect('voir_facture', facture_id=facture.pk)
        
    else:
        # Pour le GET, on redirige généralement vers la fiche client qui contient le modal
        return redirect('detail_fiche', fiche_id=fiche_id) 
    
@login_required
def listes_tapis_abandon(request):
    """
    Liste paginée des commandes de tapis abandonnées
    avec système de filtres.
    """

    STATUT_ABANDON = "ABANDON"

    # =========================
    # FILTRES (GET)
    # =========================
    search = request.GET.get("search", "")
    date_ramassage = request.GET.get("date_ramassage", "")
    nb_tapis = request.GET.get("nb_tapis", "")

    commandes = TapisDetails.objects.filter(statut=STATUT_ABANDON)

    # Recherche globale (client, numéro, localisation)
    if search:
        commandes = commandes.filter(
            Q(commande__nom_client__icontains=search) |
            Q(commande__numero_client__icontains=search) |
            Q(commande__localisation_client__icontains=search)
        )

    # Filtre date de ramassage
    if date_ramassage:
        commandes = commandes.filter(date_ramassage=date_ramassage)

    # Filtre nombre de tapis
    if nb_tapis:
        commandes = commandes.filter(nombre_tapis=nb_tapis)

    commandes = commandes.order_by("-date_ramassage")

    # =========================
    # PAGINATION
    # =========================
    paginator = Paginator(commandes, 10)  # 10 lignes par page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "commandes": page_obj,
        "page_obj": page_obj,
        "total_abandons": commandes.count(),

        # garder les valeurs des filtres
        "search": search,
        "date_ramassage": date_ramassage,
        "nb_tapis": nb_tapis,
    }

    return render(request, "index/alerte_tapis_abandon.html", context)

@login_required
@user_passes_test(is_admin)
def dashboard_financier(request):
    # 1. Base : Uniquement les factures avec optimisation de la base de données
    queryset = Facture.objects.filter(type_document='FACTURE').select_related('commande').order_by('-date_emission')

    # 2. Récupération des filtres
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    type_cmd = request.GET.get('type_commande')
    search_name = request.GET.get('search_name')

    # 3. Validation de l'intervalle de date (Sécurité Serveur)
    if start_date and end_date:
        if start_date > end_date:
            messages.error(request, "Incohérence : La date 'Du' est supérieure à la date 'Au'.")
            start_date = end_date = None  # Réinitialisation en cas d'erreur
        else:
            queryset = queryset.filter(date_emission__range=[start_date, end_date])
    elif start_date:
        queryset = queryset.filter(date_emission__gte=start_date)
    elif end_date:
        queryset = queryset.filter(date_emission__lte=end_date)

    # 4. Filtre par type et recherche par nom/numéro
    if type_cmd:
        queryset = queryset.filter(commande__type_commande=type_cmd)
    
    if search_name:
        queryset = queryset.filter(
            Q(commande__nom_client__icontains=search_name) | 
            Q(numero_document__icontains=search_name)
        )

    # 5. Calcul du montant total filtré (Le "coût" pour le client sur la période)
    total_periode = queryset.aggregate(Sum('montant_final_net'))['montant_final_net__sum'] or 0
    count_factures = queryset.count()

    # 6. Pagination compacte (15 par page)
    paginator = Paginator(queryset, 7)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'factures': page_obj,
        'total_encaisse': total_periode,
        'nombre_factures': count_factures,
        'start_date': start_date,
        'end_date': end_date,
        'type_cmd': type_cmd,
        'search_name': search_name,
        'type_choices': ['CITYPROP', 'CLIMATISEUR', 'TAPISPROP']
    }
    return render(request, 'index/dashboard_financier.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff) # Remplacez par votre test is_admin
def gestion_caisse(request):
    # 1. Gestion des dates
    maintenant = timezone.now()
    mois_id = int(request.GET.get('mois', maintenant.month))
    annee_id = int(request.GET.get('annee', maintenant.year))

    mois_noms = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin",
        7: "Juillet", 8: "Août", 9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
    }

    # 2. Calcul du Report (Solde des mois précédents)
    date_debut_mois = maintenant.replace(year=annee_id, month=mois_id, day=1, hour=0, minute=0, second=0, microsecond=0)
    mouvements_passes = OperationCaisse.objects.filter(date__lt=date_debut_mois)
    entrees_total = mouvements_passes.filter(type_mouvement='ENTREE').aggregate(Sum('montant'))['montant__sum'] or 0
    sorties_total = mouvements_passes.filter(type_mouvement='SORTIE').aggregate(Sum('montant'))['montant__sum'] or 0
    report_solde = entrees_total - sorties_total

    # 3. Préparer le Queryset filtré
    queryset_filtre = OperationCaisse.objects.filter(
        date__month=mois_id, 
        date__year=annee_id
    ).order_by('date', 'id')

    # Configuration du FormSet (extra=0 pour éviter les lignes vides fantômes)
    CaisseFormSet = modelformset_factory(
        OperationCaisse, 
        fields=('date', 'equipe', 'libelle', 'type_mouvement', 'montant'), 
        extra=0, 
        can_delete=True
    )
    
    # 4. Traitement du formulaire
    if request.method == 'POST':
        formset = CaisseFormSet(request.POST, queryset=queryset_filtre) 
        if formset.is_valid():
            formset.save() 
            messages.success(request, f"Brouillard de {mois_noms[mois_id]} mis à jour avec succès !")
            # Redirection indispensable pour éviter le message de renvoi de formulaire au rafraîchissement
            return redirect(f'/caisse/?mois={mois_id}&annee={annee_id}')
        else:
            messages.error(request, "Erreur de validation. Vérifiez que tous les champs obligatoires sont remplis.")
    else:
        formset = CaisseFormSet(queryset=queryset_filtre)
        
    context = {
        'formset': formset,
        'report_solde': report_solde,
        'mois_actuel_nom': mois_noms[mois_id],
        'mois_actuel_id': mois_id,
        'annee_actuelle': annee_id,
        'liste_mois': mois_noms,
    }
    return render(request, 'index/gestion_caisse.html', context)
# --- GENERATION EXCEL ---

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

@login_required
@user_passes_test(is_admin)
def export_caisse_excel(request, mois, annee):
    mouvements = OperationCaisse.objects.filter(date__month=mois, date__year=annee).order_by('date', 'id')
    nom_mois = dict(LISTE_MOIS).get(int(mois))

    # Calcul du report
    date_debut_mois = timezone.now().replace(year=int(annee), month=int(mois), day=1, hour=0, minute=0)
    mouvements_passes = OperationCaisse.objects.filter(date__lt=date_debut_mois)
    entrees_p = mouvements_passes.filter(type_mouvement='ENTREE').aggregate(Sum('montant'))['montant__sum'] or 0
    sorties_p = mouvements_passes.filter(type_mouvement='SORTIE').aggregate(Sum('montant'))['montant__sum'] or 0
    report_solde = entrees_p - sorties_p

    wb = Workbook()
    ws = wb.active
    ws.title = f"Brouillard {mois}-{annee}"

    # Styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid") # Orange foncé comme image
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_white = Font(color="FFFFFF", italic=True)

    # --- BLOC EN-TÊTE ADMINISTRATIF (Lignes 1 à 3) ---
    ws.merge_cells('A1:B1'); ws['A1'] = "Entreprise"
    ws.merge_cells('C1:F1'); ws['C1'] = "CITY PROP ENTRETIEN FAUTEUILLE ET TAPIS"; ws['C1'].fill = yellow_fill; ws['C1'].font = Font(bold=True)
    ws['G1'] = "N° Folio"; ws['H1'] = "Contrepartie"
    
    ws.merge_cells('A2:B2'); ws['A2'] = "Période"
    ws.merge_cells('C2:F2'); ws['C2'] = f"{nom_mois}"; ws['C2'].alignment = Alignment(horizontal="center")
    ws.merge_cells('G2:H2'); ws['G2'] = "EXERCICE COMPTABLE"
    
    ws.merge_cells('A3:B3'); ws['A3'] = "Brouillard"
    ws.merge_cells('C3:F3'); ws['C3'] = "CAISSE"; ws['C3'].alignment = Alignment(horizontal="center")
    ws.merge_cells('G3:H3'); ws['G3'] = f"{annee}"; ws['G3'].alignment = Alignment(horizontal="center")

    # Appliquer bordures au bloc d'en-tête
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=8):
        for cell in row: cell.border = thin_border

    # --- EN-TÊTES DE COLONNES (Ligne 5) ---
    headers = ['N°', 'DATE', 'Nom de chaque Equipe', 'Libellé', 'Ref', 'MONTANT ENTREE', 'MONTANT SORTIE', 'SOLDE']
    ws.append([]) # Ligne vide de séparation (Ligne 4)
    ws.append(headers) # Ligne 5
    for cell in ws[5]:
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # --- LIGNE DE REPORT (Ligne 6) ---
    ws.append(['', '', '', 'Report', '', '', '', report_solde])
    for cell in ws[6]:
        cell.fill = orange_fill
        cell.border = thin_border
        if cell.column == 4: cell.font = font_white
        if cell.column == 8: cell.font = Font(bold=True)

    # --- DONNÉES ---
    solde_courant = report_solde
    for i, m in enumerate(mouvements, 1):
        entree = m.montant if m.type_mouvement == 'ENTREE' else 0
        sortie = m.montant if m.type_mouvement == 'SORTIE' else 0
        solde_courant += (entree - sortie)

        ws.append([
            i,
            m.date.strftime("%d-%m-%y"),
            m.equipe.upper(),
            m.libelle,
            "",
            entree if entree > 0 else "",
            sortie if sortie > 0 else "",
            solde_courant
        ])
        for cell in ws[ws.max_row]: cell.border = thin_border

    # Ajustement colonnes
    column_widths = [5, 12, 30, 40, 10, 18, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Brouillard_{nom_mois}_{annee}.xlsx'
    wb.save(response)
    return response



# --- GENERATION PDF ---
@login_required
@user_passes_test(is_admin)
def export_caisse_pdf(request, mois, annee):
    # 1. Récupération des mouvements du mois
    mouvements = OperationCaisse.objects.filter(
        date__month=mois, 
        date__year=annee
    ).order_by('date', 'id')
    
    # 2. Calcul du report initial (Cumul des mois précédents)
    date_debut_mois = timezone.now().replace(year=int(annee), month=int(mois), day=1, hour=0, minute=0)
    prev = OperationCaisse.objects.filter(date__lt=date_debut_mois)
    entrees_p = prev.filter(type_mouvement='ENTREE').aggregate(Sum('montant'))['montant__sum'] or 0
    sorties_p = prev.filter(type_mouvement='SORTIE').aggregate(Sum('montant'))['montant__sum'] or 0
    report_solde = float(entrees_p - sorties_p)

    # 3. Calcul du solde progressif par ligne
    solde_courant = report_solde
    for m in mouvements:
        montant = float(m.montant) if m.montant else 0.0
        if m.type_mouvement == 'ENTREE':
            solde_courant += montant
        else:
            solde_courant -= montant
        m.solde_prog = solde_courant

    # 4. Contexte pour le template
    context = {
        'mouvements': mouvements,
        'mois_nom': dict(LISTE_MOIS).get(int(mois), "Inconnu"),
        'annee': annee,
        'report_solde': report_solde,
    }
    
    # 5. Rendu du PDF
    template = get_template('index/pdf_template.html')
    html = template.render(context)
    result = BytesIO()
    
    # xhtml2pdf génère le PDF à partir du HTML encodé
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Brouillard_{mois}_{annee}.pdf"'
        return response
    
    return HttpResponse("Erreur technique lors de la génération du PDF", status=500)

LISTE_MOIS = [
    (1, 'JAN'), (2, 'FÉV'), (3, 'MAR'), (4, 'AVR'),
    (5, 'MAI'), (6, 'JUN'), (7, 'JUL'), (8, 'AOÛ'),
    (9, 'SEP'), (10, 'OCT'), (11, 'NOV'), (12, 'DÉC')
]

@login_required
@user_passes_test(is_admin)
def gestion_caisse(request):
    try:
        mois_id = int(request.GET.get('mois', timezone.now().month))
        annee = int(request.GET.get('annee', timezone.now().year))
    except (ValueError, TypeError):
        mois_id = timezone.now().month
        annee = timezone.now().year

    queryset = OperationCaisse.objects.filter(
        date__month=mois_id, date__year=annee
    ).order_by('date', 'id')

    stats_annee = OperationCaisse.objects.filter(date__year=annee).aggregate(
        total_in=Sum('montant', filter=Q(type_mouvement='ENTREE')),
        total_out=Sum('montant', filter=Q(type_mouvement='SORTIE'))
    )
    
    solde_annuel = (stats_annee['total_in'] or 0) - (stats_annee['total_out'] or 0)

    CaisseFormSet = modelformset_factory(
        OperationCaisse,
        fields=('date', 'equipe', 'libelle', 'type_mouvement', 'montant'),
        extra=0, can_delete=True
    )

    if request.method == 'POST':
        formset = CaisseFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            instances = formset.save(commit=False)
            for obj in formset.deleted_objects:
                obj.delete()
            for instance in instances:
                if instance.date.month == mois_id and instance.date.year == annee:
                    instance.save()
            messages.success(request, "Enregistrement réussi.")
            return redirect(f"{reverse('gestion_caisse')}?mois={mois_id}&annee={annee}")

    formset = CaisseFormSet(queryset=queryset)
    
    return render(request, 'index/gestion_caisse.html', {
        'formset': formset,
        'mois_actuel_id': mois_id,
        'annee_actuelle': annee,
        'liste_mois': dict(LISTE_MOIS),
        'solde_annuel': solde_annuel,
    })
    

# Dictionnaire de correspondance (Label Excel -> Code Base de données)
TRADUCTION_STATUTS = {
    'En cours': 'NON_RESPECTE',
    'En attente': 'PRET',
    'Tapis prêt Client indisponible': 'CLIENT_INDISPO',
    'Livré - Client satisfait': 'LIVRE_SATISFAIT',
    'Livré - Client insatisfait': 'LIVRE_INSATISFAIT',
    'Tapis abandonné': 'ABANDON',
    'OK': 'OK',
    'KO_Retouche': 'KO_RET',
    'KO_Refus': 'KO_REFUS'
}

@login_required
def import_commandes_ajax(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            file = request.FILES['file']
            df = pd.read_excel(file)
            df.columns = [c.strip() for c in df.columns]
            df = df.replace({pd.NA: None, float('nan'): None}) 

            count = 0
            doublons_ignores = 0
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    line_num = index + 2 

                    # --- 1. RÉCUPÉRATION ET VALIDATION DATE (OBLIGATOIRE) ---
                    raw_date_crea = row.get('Date')
                    if not raw_date_crea:
                        raise ValueError(f"Ligne {line_num}: La 'Date' (Création) est obligatoire.")
                    
                    try:
                        date_enregistre = pd.to_datetime(raw_date_crea).date()
                    except:
                        raise ValueError(f"Ligne {line_num}: Format de Date '{raw_date_crea}' invalide.")

                    # --- 2. AUTRES PILIERS ---
                    nom = str(row.get('Nom Client') or '').strip()
                    num = str(row.get('Numéro Client') or '').strip()
                    loc = str(row.get('Localisation') or '').strip()
                    raw_type = str(row.get('Type Commande') or '').strip().upper()

                    if not all([nom, num, loc, raw_type]):
                        raise ValueError(f"Ligne {line_num}: Nom, Numéro, Localisation et Type sont obligatoires.")

                    # --- 3. GESTION DES DOUBLONS ---
                    if Commande.objects.filter(nom_client=nom, numero_client=num, 
                                            date_creation=date_enregistre, type_commande=raw_type).exists():
                        doublons_ignores += 1
                        continue

                    # --- 4. CRÉATION DE LA COMMANDE ---
                    commande = Commande.objects.create(
                        nom_client=nom, numero_client=num,
                        localisation_client=loc, type_commande=raw_type,
                        date_creation=date_enregistre
                    )

                    # --- 5. GESTION DES DÉTAILS ---
                    is_fidelise = str(row.get('Fidélisé') or '').strip().lower() in ['oui', 'yes', 'true', '1']
                    
                    def get_opt_date(val):
                        if val is None or str(val).strip() == '': return None
                        try: return pd.to_datetime(val).date()
                        except: return None

                    # Traduction du statut Excel -> Code technique
                    statut_label = str(row.get('Satisfaction / Statut') or 'En cours').strip()
                    statut_technique = TRADUCTION_STATUTS.get(statut_label, 'NON_RESPECTE')

                    if raw_type in ['CITYPROP', 'CLIMATISEUR']:
                        CityClimaDetails.objects.create(
                            commande=commande,
                            date_intervention=get_opt_date(row.get('Date Intervention')) or date_enregistre,
                            fidelise=is_fidelise,
                            satisfaction=statut_technique
                        )
                    
                    elif raw_type == 'TAPISPROP':
                        try:
                            nb = int(float(row.get('Tapis (Nb)') or 0))
                            prix = float(row.get('Coût') or 0)
                        except: nb, prix = 0, 0

                        TapisDetails.objects.create(
                            commande=commande,
                            fidelise=is_fidelise,
                            nombre_tapis=nb,
                            cout=prix,
                            date_ramassage=get_opt_date(row.get('Date Ramassage')) or date_enregistre,
                            date_traitement=get_opt_date(row.get('Date Fin Traitement')),
                            date_livraison=get_opt_date(row.get('Date Livraison')),
                            statut=statut_technique,
                            commentaire=str(row.get('Commentaire') or '').strip()
                        )
                    count += 1

            return JsonResponse({'status': 'success', 'message': f'{count} importés, {doublons_ignores} doublons ignorés.'})
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f"Erreur: {str(e)}"}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Fichier manquant.'}, status=400)

@login_required
def generer_modele_excel(request):
    colonnes = [
        'Nom Client', 'Numéro Client', 'Localisation', 'Type Commande', 'Date', 
        'Date Intervention', 'Date Ramassage', 'Date Fin Traitement', 'Date Livraison', 
        'Satisfaction / Statut', 'Fidélisé', 'Tapis (Nb)', 'Coût', 'Commentaire'
    ]
    
    # Données d'exemple avec les LABELS EXACTS du dictionnaire
    data_exemples = [
        ['Jean Dupont', '0102030405', 'Plateau', 'CITYPROP', '05/01/2026', '06/01/2026', None, None, None, 'OK', 'oui', None, 15000, 'Passage mensuel'],
        ['Marie Koné', '0708091011', 'Cocody', 'CLIMATISEUR', '05/01/2026', '07/01/2026', None, None, None, 'OK', 'non', None, 25000, 'Entretien split'],
        ['Ahmed Sylla', '0506070809', 'Marcory', 'TAPISPROP', '06/01/2026', None, '06/01/2026', '08/01/2026', '09/01/2026', 'Livré - Client satisfait', 'oui', 3, 12000, 'Tapis salon'],
        ['Moussa Diop', '0202020202', 'Treichville', 'CLIMATISEUR', '07/01/2026', '08/01/2026', None, None, None, 'KO_Refus', 'non', None, 30000, 'Rappeler demain'],
        ['Sarah Kouamé', '0303030303', 'Bingerville', 'TAPISPROP', '07/01/2026', None, '07/01/2026', None, None, 'En cours', 'non', 1, 5000, 'À récupérer'],
    ]

    df = pd.DataFrame(data_exemples, columns=colonnes)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modele_commandes.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Import')
        ws = writer.sheets['Import']

        # --- VALIDATION STATUTS (Colonne J) ---
        # Note: Les labels doivent être séparés par des virgules SANS ESPACES après la virgule
        labels_liste = "En cours,En attente,Tapis prêt Client indisponible,Livré - Client satisfait,Livré - Client insatisfait,Tapis abandonné,OK,KO_Retouche,KO_Refus"
        
        dv_sat = DataValidation(type="list", formula1=f'"{labels_liste}"', allow_blank=True)
        dv_sat.errorTitle = 'Sélection invalide'
        dv_sat.error = 'Veuillez choisir un statut dans la liste.'
        ws.add_data_validation(dv_sat)
        dv_sat.add('J2:J1000') # Applique à la colonne Satisfaction / Statut

        # --- VALIDATION FIDÉLISÉ (Colonne K) ---
        dv_fid = DataValidation(type="list", formula1='"oui,non"', allow_blank=True)
        ws.add_data_validation(dv_fid)
        dv_fid.add('K2:K1000')

        # --- VALIDATION TYPE (Colonne D) ---
        dv_type = DataValidation(type="list", formula1='"CITYPROP,CLIMATISEUR,TAPISPROP"', allow_blank=False)
        ws.add_data_validation(dv_type)
        dv_type.add('D2:D1000')

        # --- STYLE ET COULEURS ---
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in range(2, 100): # On colore les 100 premières lignes en rouge pour la date
            ws[f'E{row}'].fill = red_fill

        # Ajustement largeur
        ws.column_dimensions['J'].width = 35 # Statut (plus large pour les longs textes)
        ws.column_dimensions['A'].width = 20 # Nom
        ws.column_dimensions['D'].width = 15 # Type

    return response

def suivi_atelier_tapis(request):
    query = request.GET.get('q', '').strip()
    alerte = request.GET.get('alerte', '')
    date_filtre = request.GET.get('date', '')

    tapis_list = TapisDetails.objects.filter(
        statut__in=['NON_RESPECTE']
    ).select_related('commande')

    # 🔎 FILTRE NOM / NUMÉRO
    if query:
        tapis_list = tapis_list.filter(
            Q(commande__nom_client__icontains=query) |
            Q(commande__numero_client__icontains=query)
        )

    # 📅 FILTRE DATE (DateField)
    if date_filtre:
        tapis_list = tapis_list.filter(
            date_traitement=date_filtre
        )

    # 🚨 FILTRE PRIORITÉ (LOGIQUE IDENTIQUE À LA PROPERTY)
    if alerte:
        aujourd_hui = date.today()
        demain = aujourd_hui + timedelta(days=1)

        if alerte == "RETARD":
            tapis_list = tapis_list.filter(
                date_traitement__lt=aujourd_hui
            )

        elif alerte == "URGENT":
            tapis_list = tapis_list.filter(
                date_traitement__gte=aujourd_hui,
                date_traitement__lte=demain
            )

        elif alerte == "NORMAL":
            tapis_list = tapis_list.filter(
                Q(date_traitement__gt=demain) |
                Q(date_traitement__isnull=True)
            )

    tapis_list = tapis_list.order_by('date_traitement')

    paginator = Paginator(tapis_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'index/suivi_tapis.html', {
        'page_obj': page_obj
    })
    
    





